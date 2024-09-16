import sys
from PySide6.QtWidgets import QApplication, QDialog, QFileDialog, QMessageBox, QVBoxLayout, QPushButton, QLabel, QDateEdit
from PySide6.QtGui import QIcon
from PySide6.QtCore import QDate
from ui_untitled import Ui_Dialog
from openpyxl import load_workbook
from api import api

# Classe para o popup de data
class DateRangeDialog(QDialog):
    def __init__(self):
        super().__init__()

        # Configurando o layout do popup
        self.setWindowTitle("Selecionar Período")
        layout = QVBoxLayout(self)
        self.resize(150, 120)

        # WIdgets de data inicial
        self.label_inicial = QLabel("Data Inicial:")
        self.data_inicial = QDateEdit(self)
        self.data_inicial.setCalendarPopup(True)
        self.data_inicial.setDate(QDate.currentDate())

        # Widgets de data final
        self.label_final = QLabel("Data Final:")
        self.data_final = QDateEdit(self)
        self.data_final.setCalendarPopup(True)
        self.data_final.setDate(QDate.currentDate())

        # Botão de confirmar
        self.botao_confirmar = QPushButton("Confirmar", self)
        self.botao_confirmar.clicked.connect(self.accept)

        # Adicionando os widgets ao layout
        layout.addWidget(self.label_inicial)
        layout.addWidget(self.data_inicial)
        layout.addWidget(self.label_final)
        layout.addWidget(self.data_final)
        layout.addWidget(self.botao_confirmar)

    def get_dates(self):
        # Retorna as datas no formato 'aaaa-mm-dd'
        data_ini = self.data_inicial.date().toString("yyyy-MM-dd")
        data_fin = self.data_final.date().toString("yyyy-MM-dd")
        return data_ini, data_fin

# Classe principal
class Comparativo(QDialog):
    def __init__(self):
        super().__init__()

        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        # Ícone da janela
        self.setWindowIcon(QIcon('C:\\Users\\Luis Tvc\\Music\\Comparativo automático abastecimento\\KMD.ico'))

        # Ligação dos widgets da interface
        self.combo_combustivel = self.ui.TipoComb
        self.combo_base = self.ui.Base
        self.botao_periodo = self.ui.Periodo  
        self.botao_buscar = self.ui.BuscaExcel
        self.botao_finaliar = self.ui.Finalizar

        # Conectando o botão "Período" para abrir o popup
        self.botao_periodo.clicked.connect(self.abrir_popup_periodo)

        # Conectando outros botões
        self.botao_buscar.clicked.connect(self.selecionar_arquivo)
        self.botao_finaliar.clicked.connect(self.finalizar)

        # Inicializando as datas
        self.data_inicial = None
        self.data_final = None
        self.excel_path = None

    def abrir_popup_periodo(self):
        # Abre o popup para selecionar as datas
        popup = DateRangeDialog()
        # Se o usuário clicar em confirmar
        if popup.exec():  
            self.data_inicial, self.data_final = popup.get_dates()

    def selecionar_arquivo(self):
        # Abre um diálogo para selecionar o arquivo Excel
        caminho_arquivo, _ = QFileDialog.getOpenFileName(self, "Selecione o arquivo Excel", "", "Excel Files (*.xlsx)")

        # Se o usuário selecionou um arquivo
        if caminho_arquivo:
            self.excel_path = caminho_arquivo
        else:
            QMessageBox.warning(self, "Erro", "Nenhum arquivo foi selecionado.")

    # Função para finalizar o processo
    def finalizar(self):
        if not self.data_inicial or not self.data_final or not self.excel_path:
            QMessageBox.warning(self, "Erro", "Selecione o período e o arquivo Excel antes de finalizar.")
            return

   # Validação das datas: Verifica se a data final é posterior à data inicial
        if self.data_inicial > self.data_final:
            QMessageBox.warning(self, "Erro", "A data final não pode ser anterior à data inicial.")
            return

        # Verifica se o arquivo Excel foi selecionado
        combustivel_selecionado = self.combo_combustivel.currentText()
        base_selecionada = self.combo_base.currentText()

        # Dicionários para mapear as opções selecionadas relacionadas ao códigos do banco de dados
        base = {
            "Betim": 812,
            "São Paulo": 369,
            "Pernambuco": 923
        }

        combustivel = {
            "Diesel S10": 1,
            "Arla": 7
        }

        # Monta a consulta SQL
        query_rodopar = f"""
SELECT NUMDOC FROM RODABA 
WHERE CODPON = {base[base_selecionada]} 
AND CODCMB = {combustivel[combustivel_selecionado]} 
AND CAST(DATREF AS DATE) BETWEEN '{self.data_inicial}' AND '{self.data_final}' 
"""
        # Faz a consulta e obtém os resultados
        consulta = api("GET", query_rodopar)
        
        # Obter os números da consulta em uma lista
        numdocs_consulta = [doc['NUMDOC'] for doc in consulta]

        # Carregando o arquivo Excel
        wb = load_workbook(self.excel_path)
        ws = wb.active  # Planilha ativa

        # Excluir linhas da coluna C que não possuem "ABASTECIMENTO"
        rows_to_delete = []
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=3).value  
            # Coluna C é a coluna 3
            if cell_value and "ABASTECIMENTO" not in cell_value:
                rows_to_delete.append(row)

        # Excluir as linhas de baixo para cima (para evitar problemas ao deletar múltiplas linhas)
        for row in reversed(rows_to_delete):
            ws.delete_rows(row)

        # Comparar documentos da coluna E com a consulta (removendo 5 primeiros caracteres)
        # Lista para armazenar as linhas que serão deletadas
        rows_to_delete = [] 
        for row in range(1, ws.max_row + 1):
            # Coluna E é a coluna 5
            cell_value = ws.cell(row=row, column=5).value  
            if cell_value:
                # Remover os 5 primeiros caracteres
                doc_num = cell_value[5:]  
                if doc_num in numdocs_consulta:
                    # Adiciona a linha à lista para deletar
                    rows_to_delete.append(row)  
                else:
                    # Coluna G para resultado
                    ws.cell(row=row, column=3).value = "Não encontrado"  

        # Excluir as linhas de baixo para cima (para evitar problemas ao deletar múltiplas linhas)
        for row in reversed(rows_to_delete):
            ws.delete_rows(row)

        # Salvando o arquivo Excel com as modificações
        wb.save(self.excel_path)
        QMessageBox.information(self, "Sucesso", "As linhas foram modificadas e os resultados inseridos no Excel.")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Comparativo()
    # Exibir a janela
    window.show()
    sys.exit(app.exec())
